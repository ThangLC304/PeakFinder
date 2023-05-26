import pandas as pd
import logging
import numpy as np
import openpyxl
import os

from matplotlib.figure import Figure


logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)


## IMPORT DF FROM EXCEL ##

def init_core_folders(project_path):
    core_folders = ["Log", "Output"]
    for folder in core_folders:
        if not os.path.exists(os.path.join(project_path, folder)):
            os.makedirs(os.path.join(project_path, folder))
    

def read_clean_excel(excel_path, sheet_name=None):

    if sheet_name is None:
        # Get all sheet names of excel_path
        wb = openpyxl.load_workbook(excel_path)
        sheet_names = wb.sheetnames

        i = 0
        while True:
            sheet_name = sheet_names[i]
            if "summary" not in sheet_name.lower():
                break
            i += 1

    # load into df_whole, header = false
    df_whole = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    # if first row contain a cell with "scorer", remove it
    for cell in df_whole.iloc[0]:
        if "scorer" in str(cell).lower():
            df_whole = df_whole.drop(0)
            break

    # reset index
    df_whole = df_whole.reset_index(drop=True)

    # Combine the content of row 1 and row 2 into one row, f"{content of row 1} {content of row 2}", if content of row 2 is NaN, then just use content of row 1
    df_whole.iloc[0] = df_whole.iloc[0].fillna("")

    # iterate through cell in row 0
    for i, cell in enumerate(df_whole.iloc[0]):
        cell_below = df_whole.iloc[1][i]
        if cell_below is np.nan:
            df_whole.iloc[0][i] = cell
        else:
            df_whole.iloc[0][i] = f"{cell}_{cell_below}"

    # drop row 1
    df_whole = df_whole.drop(1)

    # make row 0 as header
    df_whole.columns = df_whole.iloc[0]

    # drop row 0
    df_whole = df_whole.drop(0)

    # reset index
    df_whole = df_whole.reset_index(drop=True)

    return df_whole

## MAKE DF FOR SAVING ##

def get_coordinates(values, positions):
    size = len(positions)
    cc = [values[positions[i]] for i in range(size)]
    return cc

def make_df(xvalues, yvalues, maxima, minima):

    xMaxima = get_coordinates(xvalues, maxima)
    yMaxima = get_coordinates(yvalues, maxima)
    xMinima = get_coordinates(xvalues, minima)
    yMinima = get_coordinates(yvalues, minima)
    
    maxima_result =  {xMaxima[i]: yMaxima[i] for i in range(len(xMaxima))}
    maxima_result = dict(sorted(maxima_result.items()))

    minima_result =  {xMinima[i]: yMinima[i] for i in range(len(xMinima))}
    minima_result = dict(sorted(minima_result.items()))

    df_max = pd.DataFrame(list(maxima_result.items()),columns = ['X_maxima', 'Y_maxima'])
    df_max['X_maxima'] = df_max['X_maxima'].astype(int)
    df_max['Y_maxima'] = df_max['Y_maxima'].astype(float)

    df_min = pd.DataFrame(list(minima_result.items()),columns = ['X_minima', 'Y_minima'])
    df_min['X_minima'] = df_min['X_minima'].astype(int)
    df_min['Y_minima'] = df_min['Y_minima'].astype(float)

    return df_max, df_min

def append_df_to_excel(filepath, df, sheet_name='Sheet1', startcol=None, startrow=None, col_sep = 0, row_sep = 0,
                       truncate_sheet=False, DISPLAY = False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filepath):
        logger.info("Excel file doesn't exist - directly export using df.to_excel")
        df.to_excel(
            filepath,
            sheet_name=sheet_name, 
            startcol=startcol if startcol is not None else 0, 
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name]
        row_0 = ws[1]
        logger.debug(f"Header: {row_0}")

        return
    
    logger.info("Excel file exists - appending using openpyxl")

    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay')

    # try to open an existing workbook
    writer.workbook = openpyxl.load_workbook(filepath)

    # get the last col in the existing Excel sheet
    # if it was not specified explicitly
    if startcol is None and sheet_name in writer.workbook.sheetnames:
        startcol = writer.workbook[sheet_name].max_column + col_sep

    if startrow is None and sheet_name in writer.workbook.sheetnames:
        startrow = writer.workbook[sheet_name].max_row + row_sep
    
    if startcol is None:
        startcol = 0

    if startrow is None:
        startrow = 0

    logger.debug(f'In file: {os.path.basename(filepath)}')
    logger.debug(f'Appended to column: {startcol}, row: {startrow}')

    try:
        row_0 = writer.workbook[sheet_name][1]
        logger.debug(f"Header: {row_0}")
    except:
        logger.debug(f"Sheet {sheet_name} doesn't exist")
    
    # remove df headers if they exist
    if startrow != 0:
        # take the first row
        first_row = df.iloc[0].astype(str)
        # check if any cell in the first row contains a letter
        has_letter = first_row.str.contains('[a-zA-Z]').any()
        if has_letter:
            df = df.iloc[1:, :]

    # write the dataframe to the existing sheet
    df.to_excel(writer, sheet_name, startcol=startcol, startrow=startrow, **to_excel_kwargs)

    # close workbook
    writer.close()

## DRAW PLOT ##

def draw_plot(xvalues, yvalues, maxima, minima):

    xMaxima = get_coordinates(xvalues, maxima)
    yMaxima = get_coordinates(yvalues, maxima)
    xMinima = get_coordinates(xvalues, minima)
    yMinima = get_coordinates(yvalues, minima)

    # Create figure and subplot
    fig = Figure(figsize=(20, 10))
    plot1 = fig.add_subplot(111)

    # Plot main data
    plot1.plot(xvalues, yvalues, label='Data')
    
    # Plot maxima points
    plot1.plot(xMaxima, yMaxima, 'ro', label='Maxima')

    # Plot minima points
    plot1.plot(xMinima, yMinima, 'bo', label='Minima')

    # Add a legend
    plot1.legend()

    return fig


